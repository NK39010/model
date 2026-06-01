# Defines input parsing for FASTA generation from manual or table input.
from __future__ import annotations

import base64
import binascii
import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from app.tools.errors import ToolInputError


SEQUENCE_ALPHABETS = {
    "dna": set("ACGTN"),
    "rna": set("ACGUN"),
    "protein": set("ABCDEFGHIKLMNPQRSTVWXYZ*"),
    "auto": set("ABCDEFGHIJKLMNOPQRSTUVWXYZ*"),
}


@dataclass(frozen=True)
class FastaRecordInput:
    name: str
    sequence: str


@dataclass(frozen=True)
class FastaGeneratorInput:
    records: list[FastaRecordInput]
    sequence_type: str = "dna"
    strict: bool = False
    wrap_length: int | None = 80

    @classmethod
    def from_payload(cls, payload: dict[str, Any]) -> "FastaGeneratorInput":
        if not isinstance(payload, dict):
            raise ToolInputError("Payload must be an object.")

        sequence_type = str(payload.get("sequence_type", "dna")).strip().lower()
        if sequence_type not in SEQUENCE_ALPHABETS:
            raise ToolInputError(
                "Unsupported sequence type.",
                {"sequence_type": sequence_type, "supported": sorted(SEQUENCE_ALPHABETS)},
            )
        strict = bool(payload.get("strict", False))
        wrap_length = _optional_bounded_int(payload.get("wrap_length", 80), 40, 200, "wrap_length")
        mode = str(payload.get("mode", "manual")).strip().lower()

        if mode == "manual":
            records = _manual_records(payload.get("records"))
        elif mode == "table":
            records = _table_records(payload)
        else:
            raise ToolInputError("Unsupported FASTA generator mode.", {"mode": mode, "supported": ["manual", "table"]})

        cleaned = _clean_records(records, sequence_type, strict)
        if not cleaned:
            raise ToolInputError("No valid FASTA records were provided.")
        return cls(records=cleaned, sequence_type=sequence_type, strict=strict, wrap_length=wrap_length)


def _manual_records(value: object) -> list[FastaRecordInput]:
    if not isinstance(value, list) or not value:
        raise ToolInputError("Manual mode requires a non-empty records list.")
    records: list[FastaRecordInput] = []
    for index, item in enumerate(value, start=1):
        if not isinstance(item, dict):
            raise ToolInputError("Each record must be an object.", {"index": index})
        records.append(FastaRecordInput(name=str(item.get("name", "")), sequence=str(item.get("sequence", ""))))
    return records


def _table_records(payload: dict[str, Any]) -> list[FastaRecordInput]:
    file_bytes, file_name = _decode_file_data_url(payload.get("file_data_url"), payload.get("file_name"))
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    has_header = str(payload.get("has_header", "auto")).strip().lower()

    if suffix == "xlsx":
        rows = _xlsx_rows(file_bytes)
    elif suffix in {"csv", "tsv", "txt"}:
        rows = _text_rows(file_bytes, suffix)
    else:
        raise ToolInputError("Unsupported table file type.", {"file_name": file_name, "supported": ["xlsx", "csv", "tsv", "txt"]})

    if has_header == "auto" and rows and _looks_like_header(rows[0]):
        rows = rows[1:]
    elif has_header == "true":
        rows = rows[1:]
    elif has_header not in {"auto", "true", "false"}:
        raise ToolInputError("has_header must be auto, true, or false.")

    records = []
    for row in rows:
        if len(row) < 2:
            continue
        name = str(row[0] or "").strip()
        sequence = str(row[1] or "").strip()
        if name or sequence:
            records.append(FastaRecordInput(name=name, sequence=sequence))
    return records


def _decode_file_data_url(value: object, file_name_value: object) -> tuple[bytes, str]:
    if not isinstance(value, str) or "," not in value:
        raise ToolInputError("Table mode requires file_data_url.")
    try:
        _, encoded = value.split(",", 1)
        file_bytes = base64.b64decode(encoded, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise ToolInputError("file_data_url contains invalid base64.") from exc
    file_name = str(file_name_value or "input.xlsx").strip() or "input.xlsx"
    return file_bytes, file_name


def _xlsx_rows(file_bytes: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ToolInputError("Could not read XLSX file.") from exc
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _text_rows(file_bytes: bytes, suffix: str) -> list[list[str]]:
    text = file_bytes.decode("utf-8-sig")
    delimiter = "\t" if suffix == "tsv" else ","
    if suffix == "txt" and "\t" in text.splitlines()[0]:
        delimiter = "\t"
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    return [row for row in reader]


def _looks_like_header(row: list[Any]) -> bool:
    if len(row) < 2:
        return False
    first = str(row[0] or "").strip().lower()
    second = str(row[1] or "").strip().lower()
    return first in {"name", "id", "名称", "名字"} and second in {"sequence", "seq", "序列"}


def _clean_records(records: list[FastaRecordInput], sequence_type: str, strict: bool) -> list[FastaRecordInput]:
    used_names: dict[str, int] = {}
    cleaned: list[FastaRecordInput] = []
    alphabet = SEQUENCE_ALPHABETS[sequence_type]

    for index, record in enumerate(records, start=1):
        name = _clean_name(record.name, index)
        sequence = "".join(str(record.sequence).split()).upper()
        invalid = sorted({char for char in sequence if char not in alphabet})
        if invalid and strict:
            raise ToolInputError("Sequence contains invalid characters.", {"name": name, "invalid": invalid})
        if invalid:
            sequence = "".join(char for char in sequence if char in alphabet)
        if not sequence:
            if strict:
                raise ToolInputError("Sequence is empty after cleaning.", {"name": name})
            continue
        name = _unique_name(name, used_names)
        cleaned.append(FastaRecordInput(name=name, sequence=sequence))
    return cleaned


def _clean_name(value: str, index: int) -> str:
    name = "_".join(str(value).strip().split())
    allowed = []
    for char in name:
        if char.isalnum() or char in {"_", "-", ".", ":"}:
            allowed.append(char)
        else:
            allowed.append("_")
    return "".join(allowed).strip("_") or f"sequence_{index}"


def _unique_name(name: str, used_names: dict[str, int]) -> str:
    count = used_names.get(name, 0)
    used_names[name] = count + 1
    if count == 0:
        return name
    return f"{name}_{count + 1}"


def _bounded_int(value: object, min_value: int, max_value: int, field_name: str) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ToolInputError(f"{field_name} must be an integer.") from exc
    if not (min_value <= parsed <= max_value):
        raise ToolInputError(
            f"{field_name} must be between {min_value} and {max_value}.",
            {field_name: parsed, "min": min_value, "max": max_value},
        )
    return parsed


def _optional_bounded_int(value: object, min_value: int, max_value: int, field_name: str) -> int | None:
    if value in (None, ""):
        return None
    return _bounded_int(value, min_value, max_value, field_name)
